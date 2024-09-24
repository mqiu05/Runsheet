import os
from dash import Dash, html, dcc, Input, Output, State
import pandas as pd
from openpyxl import load_workbook
import io

app = Dash(__name__, suppress_callback_exceptions=True)

app.layout = html.Div(children=[
    html.Button(
        'General',
        id='general-button',
        n_clicks=0,
        style={
            'outline': 'none', 'border': 'none', 'background-color': 'white',
            'color': 'black', 'padding': '10px 20px', 'font-size': '25px', 'cursor': 'pointer',
            'z-index': 2, 'text-decoration': 'underline'
        }
    ),
    html.Div(id='general-section', style={'display': 'none'}),  # Initially hidden

    html.Button(
        'Notes',
        id='notes-button',
        n_clicks=0,
        style={
            'outline': 'none', 'border': 'none', 'background-color': 'white',
            'color': 'black', 'padding': '10px 20px', 'font-size': '25px', 'cursor': 'pointer',
            'z-index': 2, 'text-decoration': 'underline'
        }
    ),
    html.Div(id='notes-section', style={'display': 'none'}),  # Initially hidden

    html.Button('Clear', id='clear-button', style={'margin-top': '20px'}),
    html.Button('Export to Excel', id='export-button', n_clicks=0),
    dcc.Download(id="download-dataframe-xlsx"),
    html.Div(id='export-status', children='Click the button to export data to Excel.')
])


@app.callback(
    Output('general-section', 'style'),
    Input('general-button', 'n_clicks'),
    State('general-section', 'style')
)
def toggle_general_section(n_clicks, style):
    if n_clicks % 2 == 1:
        return {'display': 'block'}
    return {'display': 'none'}


@app.callback(
    Output('notes-section', 'style'),
    Input('notes-button', 'n_clicks'),
    State('notes-section', 'style')
)
def toggle_notes_section(n_clicks, style):
    if n_clicks % 2 == 1:
        return {'display': 'block'}
    return {'display': 'none'}


@app.callback(
    [Output('general-section', 'children'),
     Output('notes-section', 'children')],
    [Input('general-button', 'n_clicks'),
     Input('notes-button', 'n_clicks')],
)
def update_sections(general_clicks, notes_clicks):
    general_content = html.Div([
        html.Div('Session Number', style={'grid-column': '1 / 3', 'font-weight': 'bold'}),
        dcc.Input(placeholder='Session Number', type='number', value='', id='session', style={'width': '100%'}),
        html.Div('Date', style={'grid-column': '1 / 3', 'font-weight': 'bold'}),
        dcc.DatePickerSingle(id='date-picker'),
        html.Div('Venue', style={'grid-column': '1 / 3', 'font-weight': 'bold'}),
        dcc.Dropdown(['Purdue GP', 'Venue 2', 'Venue 3'], placeholder='Venue', id='venue'),
        html.Div('Event', style={'grid-column': '1 / 3', 'font-weight': 'bold'}),
        dcc.Dropdown(['Skid', 'Brake', 'Autox'], placeholder='Event', id='event'),
        html.Div('Driver', style={'grid-column': '1 / 3', 'font-weight': 'bold'}),
        dcc.Dropdown(['Iancarlo', 'Simon', 'Troy'], placeholder='Driver', id='driver'),
        html.Div('Weight', style={'grid-column': '1 / 3', 'font-weight': 'bold'}),
        dcc.Input(placeholder='Weight', type='number', value='', id='weight', style={'width': '100%'}),
    ])

    notes_content = html.Div([
        html.Div('Driver Notes', style={'grid-column': '1 / 3', 'font-weight': 'bold'}),
        dcc.Textarea(placeholder='Driver Notes', value='', style={'width': '100%'}, id='driver-notes'),
        html.Div('Faults', style={'grid-column': '1 / 3', 'font-weight': 'bold'}),
        dcc.Textarea(placeholder='Faults', value='', style={'width': '100%'}, id='faults'),
        html.Div('Improvements', style={'grid-column': '1 / 3', 'font-weight': 'bold'}),
        dcc.Textarea(placeholder='Improvements', value='', style={'width': '100%'}, id='improvements'),
        html.Div('Misc.', style={'grid-column': '1 / 3', 'font-weight': 'bold'}),
        dcc.Textarea(placeholder='Misc. Notes', value='', style={'width': '100%'}, id='misc-notes'),
    ])

    return general_content if general_clicks else None, notes_content if notes_clicks else None


@app.callback(
    [Output('session', 'value'),
     Output('date-picker', 'date'),
     Output('venue', 'value'),
     Output('event', 'value'),
     Output('driver', 'value'),
     Output('weight', 'value'),
     Output('driver-notes', 'value'),
     Output('faults', 'value'),
     Output('improvements', 'value'),
     Output('misc-notes', 'value')],
    Input('clear-button', 'n_clicks')
)
def clear_inputs(n_clicks):
    if n_clicks:
        return '', None, '', '', '', '', '', '', '', ''


@app.callback(
    [Output('export-status', 'children'), Output('download-dataframe-xlsx', 'data')],
    Input('export-button', 'n_clicks'),
    [State('session', 'value'), State('date-picker', 'date'), State('venue', 'value'),
     State('event', 'value'), State('driver', 'value'), State('weight', 'value'),
     State('driver-notes', 'value'), State('faults', 'value'),
     State('improvements', 'value'), State('misc-notes', 'value')]
)
def export_to_excel(n_clicks, session, date, venue, event, driver, weight, driver_notes, faults, improvements,
                    misc_notes):
    if n_clicks > 0:
        # Collect data
        data = {
            'Session': [session],
            'Date': [date],
            'Venue': [venue],
            'Event': [event],
            'Driver': [driver],
            'Weight': [weight],
            'Driver Notes': [driver_notes],
            'Faults': [faults],
            'Improvements': [improvements],
            'Misc Notes': [misc_notes]
        }

        # Create DataFrame
        df = pd.DataFrame(data)

        # Define file path
        filename = 'exported_data.xlsx'
        filepath = os.path.join(os.getcwd(), filename)

        if os.path.exists(filepath):
            # If file exists, append data to it without overwriting
            book = load_workbook(filepath)
            with pd.ExcelWriter(filepath, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index=False,
                            header=False)
        else:
            # If the file doesn't exist, create a new one
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sheet1', index=False)

        # Create a downloadable version for user
        stream = io.BytesIO()
        with pd.ExcelWriter(stream, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
        stream.seek(0)

        return f'Data exported to {filename}', dcc.send_bytes(stream.getvalue(), filename=filename)

    return 'Click the button to export data to Excel.', None


if __name__ == '__main__':
    app.run_server(debug=True)
